import pandas as pd
import numpy as np
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill
from check import md_courses, md_instructor, md_rooms, md_time


# ===== BASIC ANALYSIS FUNCTIONS =====

def get_instructor_credits(df):
    """Calculate total credit hours per instructor."""
    unique_courses = df[['Instructor Name', 'Subject', 'Number', 'Section', 'Course Credit Hours']].drop_duplicates()
    return unique_courses.groupby('Instructor Name')['Course Credit Hours'].sum()


def get_instructor_schedule(df, instructor_name):
    """Get all courses for a specific instructor."""
    return df[df['Instructor Name'] == instructor_name]


def is_online_course(row):
    """Check if a course is online based on section or meeting days."""
    section = str(row['Section']).upper()
    days = str(row['Meeting Days']).strip()
    return (section.startswith(('AT', 'F', 'TC')) or 
            pd.isna(row['Meeting Days']) or 
            days == '' or days == 'nan')


# ===== CONFLICT DETECTION =====

def find_instructor_conflicts(df):
    """Find scheduling conflicts for instructors."""
    # Filter to face-to-face courses only
    face_to_face = df[~df.apply(is_online_course, axis=1) & df['Beginning Time'].notna()].copy()
    
    conflicts = []
    
    for instructor, courses in face_to_face.groupby('Instructor Name'):
        # Group by time and check for overlaps
        for time_slot, time_courses in courses.groupby('Beginning Time'):
            for day in ['M', 'T', 'W', 'R', 'F']:
                # Find courses that meet on this day at this time
                day_courses = time_courses[time_courses['Meeting Days'].str.contains(day, na=False)]
                
                if len(day_courses) > 1:
                    conflicts.append({
                        'instructor': instructor,
                        'time': int(time_slot),
                        'day': day,
                        'courses': day_courses[['Subject', 'Number', 'Section', 'Catalog Title']].to_dict('records')
                    })
    
    return conflicts


def find_room_conflicts(df):
    """Find scheduling conflicts for rooms."""
    # Filter to face-to-face courses with room info
    with_rooms = df[~df.apply(is_online_course, axis=1) & 
                   df['Beginning Time'].notna() & 
                   df['Building'].notna() & 
                   df['Room'].notna()].copy()
    
    with_rooms['room_key'] = with_rooms['Building'].str.strip() + ' ' + with_rooms['Room'].astype(str).str.strip()
    
    conflicts = []
    
    for room, courses in with_rooms.groupby('room_key'):
        for time_slot, time_courses in courses.groupby('Beginning Time'):
            for day in ['M', 'T', 'W', 'R', 'F']:
                day_courses = time_courses[time_courses['Meeting Days'].str.contains(day, na=False)]
                
                if len(day_courses) > 1:
                    conflicts.append({
                        'room': room,
                        'time': int(time_slot),
                        'day': day,
                        'courses': day_courses[['Subject', 'Number', 'Section', 'Instructor Name']].to_dict('records')
                    })
    
    return conflicts


def print_conflicts(conflicts, conflict_type="instructor"):
    """Print conflicts in a readable format."""
    if not conflicts:
        print(f"✅ No {conflict_type} conflicts found!")
        return
    
    print(f"⚠️  {conflict_type.upper()} CONFLICTS DETECTED")
    print("=" * 50)
    
    for i, conflict in enumerate(conflicts, 1):
        key = conflict.get('instructor', conflict.get('room'))
        print(f"\nConflict #{i}: {key}")
        print(f"Time: {conflict['day']} at {conflict['time']}")
        print("Overlapping courses:")
        
        for course in conflict['courses']:
            if conflict_type == "instructor":
                print(f"  - {course['Subject']}{course['Number']}-{course['Section']}: {course['Catalog Title']}")
            else:
                print(f"  - {course['Subject']}{course['Number']}-{course['Section']}: {course['Instructor Name']}")


# ===== SECTION COUNTING =====

def count_sections(df):
    """Count sections by course, separating online and in-person."""
    df['course_code'] = df['Subject'] + ' ' + df['Number'].astype(str)
    df['is_online'] = df.apply(is_online_course, axis=1)
    
    summary = df.groupby('course_code').agg({
        'is_online': ['sum', 'count']
    }).reset_index()
    
    summary.columns = ['Course', 'Online_Sections', 'Total_Sections']
    summary['In_Person_Sections'] = summary['Total_Sections'] - summary['Online_Sections']
    
    return summary[['Course', 'In_Person_Sections', 'Online_Sections', 'Total_Sections']].sort_values('Course')


def print_section_summary(df):
    """Print a formatted summary of course sections."""
    summary = count_sections(df)
    
    print("\nCOURSE SECTION SUMMARY")
    print("=" * 60)
    print(f"{'Course':<12} {'In-Person':>10} {'Online':>8} {'Total':>8}")
    print("-" * 60)
    
    for _, row in summary.iterrows():
        print(f"{row['Course']:<12} {row['In_Person_Sections']:>10} {row['Online_Sections']:>8} {row['Total_Sections']:>8}")
    
    print("=" * 60)
    print(f"TOTALS: {summary['In_Person_Sections'].sum()} in-person, {summary['Online_Sections'].sum()} online")


# ===== EXCEL GENERATION =====

def format_time(time_val):
    """Convert 930.0 to '9:30'."""
    if pd.isna(time_val):
        return ''
    hours = int(time_val // 100)
    minutes = int(time_val % 100)
    return f"{hours}:{minutes:02d}"


def create_instructor_schedule_excel(df, output_path="instructor_schedules.xlsx"):
    """Create Excel file with instructor schedules."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    for instructor in df['Instructor Name'].unique():
        instructor_courses = df[df['Instructor Name'] == instructor]
        
        # Create sheet for instructor
        sheet_name = instructor[:30]  # Excel sheet name limit
        ws = wb.create_sheet(sheet_name)
        
        # Headers
        headers = ['Course', 'Section', 'Title', 'Days', 'Time', 'Room', 'Credits']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Course data
        for row, course in enumerate(instructor_courses.iterrows(), 2):
            course_data = course[1]
            ws.cell(row=row, column=1, value=f"{course_data['Subject']} {course_data['Number']}")
            ws.cell(row=row, column=2, value=course_data['Section'])
            ws.cell(row=row, column=3, value=course_data['Catalog Title'])
            ws.cell(row=row, column=4, value=course_data['Meeting Days'] or 'Online')
            
            if pd.notna(course_data['Beginning Time']):
                time_str = f"{format_time(course_data['Beginning Time'])}-{format_time(course_data['Ending Time'])}"
            else:
                time_str = 'Online'
            ws.cell(row=row, column=5, value=time_str)
            
            room = ''
            if pd.notna(course_data['Building']) and pd.notna(course_data['Room']):
                room = f"{course_data['Building']} {course_data['Room']}"
            ws.cell(row=row, column=6, value=room)
            
            ws.cell(row=row, column=7, value=course_data['Course Credit Hours'])
    
    wb.save(output_path)
    print(f"Instructor schedules saved to {output_path}")


# ===== MARKDOWN EXPORT =====

def create_schedule_markdown(df, output_path="schedule.md"):
    """Create markdown file with complete schedule."""
    lines = ["# Course Schedule\n"]
    
    # Group by course
    df['course_key'] = df['Subject'] + ' ' + df['Number'].astype(str)
    
    for course, course_sections in df.groupby('course_key'):
        title = course_sections.iloc[0]['Catalog Title']
        lines.append(f"## {course} - {title}\n")
        lines.append("| Section | Instructor | Days | Time | Room | Credits |")
        lines.append("|---------|------------|------|------|------|---------|")
        
        for _, section in course_sections.iterrows():
            days = section['Meeting Days'] if pd.notna(section['Meeting Days']) else 'Online'
            
            if pd.notna(section['Beginning Time']):
                time_str = f"{format_time(section['Beginning Time'])}-{format_time(section['Ending Time'])}"
            else:
                time_str = 'Online'
            
            room = ''
            if pd.notna(section['Building']) and pd.notna(section['Room']):
                room = f"{section['Building']} {section['Room']}"
            
            lines.append(f"| {section['Section']} | {section['Instructor Name']} | {days} | {time_str} | {room} | {section['Course Credit Hours']} |")
        
        lines.append("")
    
    with open(output_path, 'w') as f:
        f.write('\n'.join(lines))
    
    print(f"Schedule saved to {output_path}")


# ===== MAIN ANALYSIS FUNCTION =====

def analyze_schedule(df, output_folder="analysis"):
    """Run complete schedule analysis."""
    Path(output_folder).mkdir(exist_ok=True)
    
    print("SCHEDULE ANALYSIS")
    print("=" * 50)
    
    # Check conflicts
    instructor_conflicts = find_instructor_conflicts(df)
    print_conflicts(instructor_conflicts, "instructor")
    
    room_conflicts = find_room_conflicts(df)
    print_conflicts(room_conflicts, "room")
    
    # Section summary
    print_section_summary(df)
    
    # Credit hours by instructor
    credits = get_instructor_credits(df)
    print(f"\nINSTRUCTOR CREDIT HOURS")
    print("=" * 30)
    for instructor, total_credits in credits.sort_values(ascending=False).items():
        print(f"{instructor}: {total_credits} credits")
    
    # Generate files
    create_instructor_schedule_excel(df, Path(output_folder) / "instructor_schedules.xlsx")
    # create_course_based_markdown(df, Path(output_folder) / "courses.md")
    # create_instructor_based_markdown(df, Path(output_folder) / "instructors.md")
    # create_room_based_markdown(df, Path(output_folder) / "rooms.md")
    # create_time_based_markdown(df, Path(output_folder) / "timeslots.md")
    
    # Save raw data
    df.to_excel(Path(output_folder) / "raw_schedule.xlsx", index=False)
    
    return {
        'instructor_conflicts': instructor_conflicts,
        'room_conflicts': room_conflicts,
        'section_counts': count_sections(df),
        'credit_hours': credits
    }


# ===== USAGE EXAMPLE =====
# df = pd.read_csv('your_schedule.csv')
# results = analyze_schedule(df)