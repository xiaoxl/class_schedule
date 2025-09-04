import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill
# from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from datetime import datetime


# from planner import show_instructor_schedule
from conflicts import (
    check_instructor_conflicts_matrix,
    print_instructor_matrix_conflicts,
    check_room_conflicts_matrix,
    print_room_matrix_conflicts,
)

daymapping = {
    "B2": "Monday",
    "C2": "Tuesday",
    "D2": "Wednesday",
    "E2": "Thursday",
    "F2": "Friday",
}

reversedaymapping = {"M": "B", "T": "C", "W": "D", "R": "E", "F": "F"}

timemapping = {
    800: '8:00-8:30',
    830: '8:30-9:00',
    900: '9:00-9:30',
    930: '9:30-10:00',
    1000: '10:00-10:30',
    1030: '10:30-11:00',
    1100: '11:00-11:30',
    1130: '11:30-12:00',
    1200: '12:00-12:30',
    1230: '12:30-13:00',
    1300: '13:00-13:30',
    1330: '13:30-14:00',
    1400: '14:00-14:30',
    1430: '14:30-15:00',
    1500: '15:00-15:30',
    1530: '15:30-16:00',
    1600: '16:00-16:30',
    1630: '16:30-17:00',
}



timemapping = {datetime.strptime(str(k).zfill(4), '%H%M').time(): v for k, v in timemapping.items()}

# daymapping = {
#     'B2': 'Monday',
#     'C2': 'Tuesday',
#     'D2': 'Wednesday',
#     'E2': 'Thursday',
#     'F2': 'Friday',
# }


def time2idx(t):
    # idx = int((t-800)//100*2+np.floor(((t-800)%100)/30))
    # idx = int(((t - 800) // 100) * 2 + ((t % 100) // 30))
    minutes_since_8 = (t.hour - 8) * 60 + t.minute
    return minutes_since_8 // 30
    # return idx


# def idx2time(idx):
#     i = idx//2
#     j = idx%2
#     return 800+100*i+j*30


def generate_table(st):
    for k, v in daymapping.items():
        st[k] = v
        st[k].alignment = Alignment(horizontal="center",
                                    vertical="center")
        st.column_dimensions[k[0]].width = 15
    for k, v in timemapping.items():
        idx = time2idx(k)
        st[f"A{idx + 3}"] = v
        st[f"A{idx + 3}"].alignment = Alignment(horizontal="center",
                                                vertical="center")
        st.column_dimensions["A"].width = 15
    st.merge_cells('A21:A22')
    st['A21'].alignment = Alignment(wrap_text=True,
                                    horizontal="center",
                                    vertical="center")
    st['A21'] = 'Online'
    return st


BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

COLORS = [
    "E6F3FF",  # Light blue
    "E6FFE6",  # Light green
    "FFF2E6",  # Light orange
    "F0E6FF",  # Light purple
    "FFE6F2",  # Light pink
    "E6FFFF",  # Light cyan
    "FFFACD",  # Light yellow
    "F5F5DC",  # Light beige
    "E6E6FA",  # Light lavender
    "F0F8E6",  # Light mint
]


def add_a_course_to_a_cell(st, cells, text, color=0, border=BORDER, colors=COLORS):
    st.merge_cells(cells)
    firstcell = cells.split(":")[0]
    st[firstcell] = text
    st[firstcell].alignment = Alignment(wrap_text=True,
                                        horizontal="center",
                                        vertical="center")
    st[firstcell].border = border
    for row in st[cells]:
        for cell in row:
            cell.border = border
    colname = firstcell[0]
    firstrow = int(firstcell[1:])
    lastrow = int(cells.split(":")[1][1:])
    st.column_dimensions[colname].width = 15
    for i in range(firstrow, lastrow + 1):
        st.row_dimensions[i].height = 30
    # print(color)
    st[firstcell].fill = PatternFill(start_color=colors[color], end_color=colors[color], fill_type="solid")
    return st


def add_a_row(st, row, color=0):
    instructor_text = f"{row['Subject']} {row['Number']} {row['Section']} {row['Room']}"
    # if pd.isna(row["Building"]) or pd.isna(row["Room"]):
    #     room_text = ""
    # else:
    #     room_text = f"{row['Building']} {int(row['Room'])}"
    # instructor_text = f"{course_text} {room_text}"
    if not pd.isna(row["Meeting Days"]):
        for day in row["Meeting Days"]:
            col = reversedaymapping[day]
            r1 = time2idx(row["Beginning Time"])
            r2 = time2idx(row["Ending Time"])
            cells = f"{col}{r1 + 3}:{col}{r2 + 3}"
            add_a_course_to_a_cell(st, cells, instructor_text, color)
    else:
        rnum = 21
        cnum = 2
        found = False
        while not found:
            c_cell = st.cell(row=rnum, column=cnum).value
            if c_cell is not None:
                cnum += 1
            else:
                found = True


        st.merge_cells(f"{get_column_letter(cnum)}{rnum}:{get_column_letter(cnum)}{rnum+1}")
        f_r = f"{get_column_letter(cnum)}{rnum}"

        st[f_r] = instructor_text
        st[f_r].alignment = Alignment(wrap_text=True,
                                      horizontal="center",
                                      vertical="center")
        st[f_r].fill = PatternFill(start_color=COLORS[color],
                                   end_color=COLORS[color],
                                   fill_type="solid")
        for i in range(rnum, rnum + 2):
            st.row_dimensions[i].height = 30
            st[f"{get_column_letter(cnum)}{i}"].border = BORDER




        # rownum = 18
        
        # found = False
        # while not found:
        #     cell = st[f"B{rownum}"]
        #     if not isinstance(cell, MergedCell) and cell.value is None:
        #         found = True
        #     else:
        #         rownum += 1
        # st.merge_cells(f"B{rownum}:B{rownum + 1}")
        # st[f"B{rownum}"] = instructor_text
        # st[f"B{rownum}"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        # st[f"B{rownum}"].border = BORDER
        # # for row in st[cells]:
        # #     for cell in row:
        # #         cell.border = BORDER
        # for i in range(rownum, rownum + 2):
        #     st.row_dimensions[i].height = 30
        # st[f"B{rownum}"].fill = PatternFill(start_color=COLORS[color], end_color=COLORS[color], fill_type="solid")
    return st


def add_same_instructors(st, df, name):
    # sf = df.dropna()
    if not pd.isna(name):
        sf = df[df["Instructor Name"] == name]  # .dropna()
    else:
        sf = df[df["Instructor Name"].isna()]
    # print(sf)
    color_idx = {
        f"{r['Number']} {r['Section']}": i
        for i, r in sf[["Subject", "Number", "Section"]].drop_duplicates().reset_index().iterrows()
    }
    for i, row in sf.iterrows():
        add_a_row(st, row, color_idx[f"{row['Number']} {row['Section']}"])
    return st


def add_a_row_room(st, row, color=0):
    instructor_text = f"{row['Subject']} {row['Number']} {row['Section']} {row['Instructor Name']}"
    for day in row["Meeting Days"]:
        col = reversedaymapping[day]
        r1 = time2idx(row["Beginning Time"])
        r2 = time2idx(row["Ending Time"])
        cells = f"{col}{r1 + 3}:{col}{r2 + 3}"
        add_a_course_to_a_cell(st, cells, instructor_text, color)
    return st


def add_same_room(st, df, room):
    sf = df.dropna()
    sf = sf[sf["Room"] == room]
    # print(sf)
    color_idx = {
        f"{r['Number']} {r['Section']}": i
        for i, r in sf[["Subject", "Number", "Section"]].drop_duplicates().reset_index().iterrows()
    }
    for i, row in sf.iterrows():
        add_a_row_room(st, row, color_idx[f"{row['Number']} {row['Section']}"])
    return st


def room_excel(wb, df):
    rooms = df[["Room"]].drop_duplicates().dropna()
    for i, r in rooms.iterrows():
        st = wb.create_sheet(f"{r['Room']}")
        generate_table(st)
        add_same_room(st, df, r["Room"])


def instructor_excel(wb, df):
    names = df["Instructor Name"].drop_duplicates()
    for name in names:
        st = wb.create_sheet(f'{name}')
        generate_table(st)
        add_same_instructors(st, df, name)


def md_instructor(df):
    t = ""
    for (i), r in df.groupby(["Instructor Name"]):
        t += f"## {i[0]}\n"
        t += "| Course | Section | Days | Time | Room |\n"
        t += "|---------|------------|------|------|------|\n"
        for _, row in r.iterrows():
            course = f"{row['Subject']} {row['Number']}"
            section = row['Section']
            days = "Online" if pd.isna(row['Meeting Days']) else row['Meeting Days']
            time = "" if pd.isna(row['Beginning Time']) else convert_time(row['Beginning Time'])
            room = "" if pd.isna(row['Room']) else f"{row['Room']}"

            t += f"| {course} | {section} | {days} | {time} | {room} |\n"
    return t


def print_instructor(df, folder="out"):
    with open(Path(folder) / "instructor.md", "w") as f:
        f.write(md_instructor(df))

def convert_time(b):
    return b
    # return f"{int(b//100)}:{int(b)%100:02d}"


def md_time(df):
    t = ""
    for (m, b), r in df.groupby(["Meeting Days", "Beginning Time"]):
        t += f"# {m} {convert_time(b)}\n"
        t += "| Course | Section | Instructor |  Room |\n"
        t += "|---------|------------|------|------|\n"
        for _, row in r.iterrows():
            course = f"{row['Subject']} {row['Number']}"
            section = row['Section']
            name = row['Instructor Name']
            # days = "Online" if pd.isna(row['Meeting Days']) else row['Meeting Days']
            # time = "" if pd.isna(row['Beginning Time']) else row['Beginning Time']
            room = "Online" if pd.isna(row['Room']) else f"{row['Room']}"

            t += f"| {course} | {section} | {name} | {room} |\n"
    return t


def print_time_slots(df, folder="out"):
    with open(Path(folder) / "time_slot.md", "w") as f:
        f.write(md_time(df))
        # for (m, b), d in df.groupby(["Meeting Days", "Beginning Time"]):
        #     f.write(f"# {m} {b}\n")
        #     for j, r in d.iterrows():
        #         f.write(
        #             f"- {r['Subject']} {r['Number']}-{r['Section']}: {r['Instructor Name']} {r['Building']} {r['Room']}\n"
        #         )
        #     f.write("\n")


def md_courses(df):
    t = ""
    for (m, b), r in df.groupby(["Subject", "Number"]):
        t += f"# {m} {b}\n"
        t += "| Section | Instructor | Days | Time | Room |\n"
        t += "|---------|------------|------|------|------|\n"
        for _, row in r.iterrows():
            # course = f"{row['Subject']} {row['Number']}"
            section = row['Section']
            name = row['Instructor Name']
            days = "Online" if pd.isna(row['Meeting Days']) else row['Meeting Days']
            time = "" if pd.isna(row['Beginning Time']) else convert_time(row['Beginning Time'])
            room = "Online" if pd.isna(row['Room']) else f"{row['Room']}"

            t += f"| {section} | {name} | {days} | {time} | {room} |\n"
    return t


def print_courses(df, folder="out"):
    with open(Path(folder) / "courses.md", "w") as f:
        f.write(md_courses(df))
        # for (m, b), d in df.groupby(["Subject", "Number"]):
        #     f.write(f"# {m} {b}\n")
        #     for j, r in d.iterrows():
        #         f.write(
        #             f"- {r['Section']}: {r['Instructor Name']} {r['Meeting Days']} {r['Beginning Time']} {r['Building']} {r['Room']}\n"
        #         )
        #     f.write("\n")


def md_rooms(df):
    t = ""
    for (b), r in df.groupby(["Room"]):
        if not pd.isna(b):
            t += f"# {b[0]}\n"
            t += "| Days | Time | Course | Section | Instructor |\n"
            t += "|---------|------------|------|------|------|\n"
            for _, row in r.iterrows():
                course = f"{row['Subject']} {row['Number']}"
                section = row['Section']
                name = row['Instructor Name']
                days = "Online" if pd.isna(row['Meeting Days']) else row['Meeting Days']
                time = "" if pd.isna(row['Beginning Time']) else convert_time(row['Beginning Time'])
                # room = "Online" if pd.isna(row['Building']) else f"{row['Building']} {int(row['Room'])}"

                t += f"| {days} | {time} | {course} | {section} | {name} |\n"
    return t


def print_room(df, folder="out"):
    with open(Path(folder) / "rooms.md", "w") as f:
        f.write(md_rooms(df))
        # for (m, b), d in df.groupby(["Building", "Room"]):
        #     f.write(f"# {m} {b}\n")
        #     for j, r in d.iterrows():
        #         f.write(
        #             f"- {r['Meeting Days']} {r['Beginning Time']}: {r['Subject']} {r['Number']}-{r['Section']} {r['Instructor Name']} \n"
        #         )
        #     f.write("\n")

def md_compute_credits(df):
    credit_clean = (
        df[["Instructor Name", "Subject", "Number", "Section", "Credits"]]
        .drop_duplicates()
        .reset_index(drop=True)
    )
    d = credit_clean[["Credits"]].groupby([credit_clean["Instructor Name"]]).sum().to_dict()['Credits']

    t = "| Instructor | Credits |\n"
    t += "|---------|------------|\n"
    for k, v in d.items():
        t += f"| {k} | {v} |\n"
    return t

def compute_credits(df):
    credit_clean = (
        df[["Instructor Name", "Subject", "Number", "Section", "Credits"]]
        .drop_duplicates()
        .reset_index(drop=True)
    )
    return credit_clean[["Credits"]].groupby([credit_clean["Instructor Name"]]).sum()


def whole_package(df, folder="out"):
    Path(folder).mkdir(parents=True, exist_ok=True)

    instructor_conflicts = check_instructor_conflicts_matrix(df)
    print_instructor_matrix_conflicts(instructor_conflicts)

    room_conflicts = check_room_conflicts_matrix(df)
    print_room_matrix_conflicts(room_conflicts)

    # print(compare_instructors(sf, df))

    print_time_slots(df, folder)
    print_courses(df, folder)
    print_instructor(df, folder)
    print_room(df, folder)
    print(compute_credits(df))

    if (not instructor_conflicts) and (not room_conflicts):
        wbr = openpyxl.Workbook()
        wbr.remove(wbr.active)
        room_excel(wbr, df)
        wbr.save(Path(folder) / "schedule_room.xlsx")

        wbu = openpyxl.Workbook()
        wbu.remove(wbu.active)
        instructor_excel(wbu, df)
        wbu.save(Path(folder) / "schedule_instructor.xlsx")

    df.to_excel(Path(folder) / "schedule.xlsx", index=False)
    # return wbr
