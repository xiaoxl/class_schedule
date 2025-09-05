import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from IPython.display import Markdown, display


from conflicts import (
    check_instructor_conflicts_matrix,
    md_instructor_matrix_conflicts,
    check_room_conflicts_matrix,
    md_room_matrix_conflicts,
)

daymapping = {
    "B2": "Monday",
    "C2": "Tuesday",
    "D2": "Wednesday",
    "E2": "Thursday",
    "F2": "Friday",
}

reversedaymapping = {"M": "B", "T": "C", "W": "D", "R": "E", "F": "F"}

timemapping = {
    800: '8:00-8:30',
    830: '8:30-9:00',
    900: '9:00-9:30',
    930: '9:30-10:00',
    1000: '10:00-10:30',
    1030: '10:30-11:00',
    1100: '11:00-11:30',
    1130: '11:30-12:00',
    1200: '12:00-12:30',
    1230: '12:30-13:00',
    1300: '13:00-13:30',
    1330: '13:30-14:00',
    1400: '14:00-14:30',
    1430: '14:30-15:00',
    1500: '15:00-15:30',
    1530: '15:30-16:00',
    1600: '16:00-16:30',
    1630: '16:30-17:00',
}

timemapping = {datetime.strptime(str(k).zfill(4), '%H%M').time(): v for k, v in timemapping.items()}


BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

COLORS = [
    "E6F3FF",  # Light blue
    "E6FFE6",  # Light green
    "FFF2E6",  # Light orange
    "F0E6FF",  # Light purple
    "FFE6F2",  # Light pink
    "E6FFFF",  # Light cyan
    "FFFACD",  # Light yellow
    "F5F5DC",  # Light beige
    "E6E6FA",  # Light lavender
    "F0F8E6",  # Light mint
]


def time2idx(t):
    minutes_since_8 = (t.hour - 8) * 60 + t.minute
    return minutes_since_8 // 30


def generate_table(st):
    for k, v in daymapping.items():
        st[k] = v
        st[k].alignment = Alignment(horizontal="center",
                                    vertical="center")
        st.column_dimensions[k[0]].width = 15
    for k, v in timemapping.items():
        idx = time2idx(k)
        st[f"A{idx + 3}"] = v
        st[f"A{idx + 3}"].alignment = Alignment(horizontal="center",
                                                vertical="center")
        st.column_dimensions["A"].width = 15
    st.merge_cells('A21:A22')
    st['A21'].alignment = Alignment(wrap_text=True,
                                    horizontal="center",
                                    vertical="center")
    st['A21'] = 'Online'
    return st


def add_a_course_to_a_cell(st, cells, text, color=0, border=BORDER, colors=COLORS):
    st.merge_cells(cells)
    firstcell = cells.split(":")[0]
    st[firstcell] = text
    st[firstcell].alignment = Alignment(wrap_text=True,
                                        horizontal="center",
                                        vertical="center")
    st[firstcell].border = border
    for row in st[cells]:
        for cell in row:
            cell.border = border
    colname = firstcell[0]
    firstrow = int(firstcell[1:])
    lastrow = int(cells.split(":")[1][1:])
    st.column_dimensions[colname].width = 15
    for i in range(firstrow, lastrow + 1):
        st.row_dimensions[i].height = 30
    st[firstcell].fill = PatternFill(start_color=colors[color], end_color=colors[color], fill_type="solid")
    return st


def add_a_row(st, row, color=0):
    instructor_text = f"{row['Subject']} {row['Number']} {row['Section']} {row['Room']}"
    if not pd.isna(row["Meeting Days"]):
        for day in row["Meeting Days"]:
            col = reversedaymapping[day]
            r1 = time2idx(row["Beginning Time"])
            r2 = time2idx(row["Ending Time"])
            cells = f"{col}{r1 + 3}:{col}{r2 + 3}"
            add_a_course_to_a_cell(st, cells, instructor_text, color)
    else:
        rnum = 21
        cnum = 2
        found = False
        while not found:
            c_cell = st.cell(row=rnum, column=cnum).value
            if c_cell is not None:
                cnum += 1
            else:
                found = True

        st.merge_cells(f"{get_column_letter(cnum)}{rnum}:{get_column_letter(cnum)}{rnum+1}")
        f_r = f"{get_column_letter(cnum)}{rnum}"

        st[f_r] = instructor_text
        st[f_r].alignment = Alignment(wrap_text=True,
                                      horizontal="center",
                                      vertical="center")
        st[f_r].fill = PatternFill(start_color=COLORS[color],
                                   end_color=COLORS[color],
                                   fill_type="solid")
        for i in range(rnum, rnum + 2):
            st.row_dimensions[i].height = 30
            st[f"{get_column_letter(cnum)}{i}"].border = BORDER

    return st


def add_same_instructors(st, df, name):
    if not pd.isna(name):
        sf = df[df["Instructor Name"] == name] 
    else:
        sf = df[df["Instructor Name"].isna()]
    color_idx = {
        f"{r['Number']} {r['Section']}": i
        for i, r in sf[["Subject", "Number", "Section"]].drop_duplicates().reset_index().iterrows()
    }
    for i, row in sf.iterrows():
        add_a_row(st, row, color_idx[f"{row['Number']} {row['Section']}"])
    return st


def add_a_row_room(st, row, color=0):
    instructor_text = f"{row['Subject']} {row['Number']} {row['Section']} {row['Instructor Name']}"
    for day in row["Meeting Days"]:
        col = reversedaymapping[day]
        r1 = time2idx(row["Beginning Time"])
        r2 = time2idx(row["Ending Time"])
        cells = f"{col}{r1 + 3}:{col}{r2 + 3}"
        add_a_course_to_a_cell(st, cells, instructor_text, color)
    return st


def add_same_room(st, df, room):
    sf = df.dropna()
    sf = sf[sf["Room"] == room]
    color_idx = {
        f"{r['Number']} {r['Section']}": i
        for i, r in sf[["Subject", "Number", "Section"]].drop_duplicates().reset_index().iterrows()
    }
    for i, row in sf.iterrows():
        add_a_row_room(st, row, color_idx[f"{row['Number']} {row['Section']}"])
    return st


def room_excel(wb, df):
    rooms = df[["Room"]].drop_duplicates().dropna()
    for i, r in rooms.iterrows():
        st = wb.create_sheet(f"{r['Room']}")
        generate_table(st)
        add_same_room(st, df, r["Room"])


def instructor_excel(wb, df):
    names = df["Instructor Name"].drop_duplicates()
    for name in names:
        st = wb.create_sheet(f'{name}')
        generate_table(st)
        add_same_instructors(st, df, name)


def md_instructor(df):
    t = ""
    for (i), r in df.groupby(["Instructor Name"]):
        t += f"## {i[0]}\n"
        t += "| Course | Section | Days | Time | Room |\n"
        t += "|---------|------------|------|------|------|\n"
        for _, row in r.iterrows():
            course = f"{row['Subject']} {row['Number']}"
            section = row['Section']
            days = "Online" if pd.isna(row['Meeting Days']) else row['Meeting Days']
            time = "" if pd.isna(row['Beginning Time']) else row['Beginning Time']
            room = "" if pd.isna(row['Room']) else f"{row['Room']}"

            t += f"| {course} | {section} | {days} | {time} | {room} |\n"
    return t


def md_time(df):
    t = ""
    for (m, b), r in df.groupby(["Meeting Days", "Beginning Time"]):
        t += f"# {m} {b}\n"
        t += "| Course | Section | Instructor |  Room |\n"
        t += "|---------|------------|------|------|\n"
        for _, row in r.iterrows():
            course = f"{row['Subject']} {row['Number']}"
            section = row['Section']
            name = row['Instructor Name']
            room = "Online" if pd.isna(row['Room']) else f"{row['Room']}"

            t += f"| {course} | {section} | {name} | {room} |\n"
    return t


def md_courses(df):
    t = ""
    for (m, b), r in df.groupby(["Subject", "Number"]):
        t += f"# {m} {b}\n"
        t += "| Section | Instructor | Days | Time | Room |\n"
        t += "|---------|------------|------|------|------|\n"
        for _, row in r.iterrows():
            section = row['Section']
            name = row['Instructor Name']
            days = "Online" if pd.isna(row['Meeting Days']) else row['Meeting Days']
            time = "" if pd.isna(row['Beginning Time']) else row['Beginning Time']
            room = "Online" if pd.isna(row['Room']) else f"{row['Room']}"

            t += f"| {section} | {name} | {days} | {time} | {room} |\n"
    return t


def md_rooms(df):
    t = ""
    for (b), r in df.groupby(["Room"]):
        if not pd.isna(b):
            t += f"# {b[0]}\n"
            t += "| Days | Time | Course | Section | Instructor |\n"
            t += "|---------|------------|------|------|------|\n"
            for _, row in r.iterrows():
                course = f"{row['Subject']} {row['Number']}"
                section = row['Section']
                name = row['Instructor Name']
                days = "Online" if pd.isna(row['Meeting Days']) else row['Meeting Days']
                time = "" if pd.isna(row['Beginning Time']) else row['Beginning Time']
                t += f"| {days} | {time} | {course} | {section} | {name} |\n"
    return t


def md_compute_credits(df):
    credit_clean = (
        df[["Instructor Name", "Subject", "Number", "Section", "Credits"]]
        .drop_duplicates()
        .reset_index(drop=True)
    )
    d = credit_clean[["Credits"]].groupby([credit_clean["Instructor Name"]]).sum().to_dict()['Credits']

    t = "| Instructor | Credits |\n"
    t += "|---------|------------|\n"
    for k, v in d.items():
        t += f"| {k} | {v} |\n"
    return t


def generate_reports(df):
    instructor_conflicts = check_instructor_conflicts_matrix(df)
    ic = md_instructor_matrix_conflicts(instructor_conflicts)
    room_conflicts = check_room_conflicts_matrix(df)
    rc = md_room_matrix_conflicts(room_conflicts)
    t = md_time(df)
    n = md_instructor(df)
    c = md_courses(df)
    r = md_rooms(df)
    h = md_compute_credits(df)
    return {
        'instructor_conflicts': instructor_conflicts,
        'room_conflicts': room_conflicts,
        'instructor_conflicts_text': ic,
        'room_conflicts_text': rc,
        'schedule_time': t,
        'schedule_instructor': n,
        'schedule_course': c,
        'schedule_room': r,
        'instructor_credits': h
    }


def save_reports(df, folder="out"):
    Path(folder).mkdir(parents=True, exist_ok=True)

    reports = generate_reports(df)

    instructor_conflicts = reports['instructor_conflicts']
    ic = reports['instructor_conflicts_text']
    room_conflicts = reports['room_conflicts']
    rc = reports['room_conflicts_text']
    t = reports['schedule_time']
    n = reports['schedule_instructor']
    c = reports['schedule_course']
    r = reports['schedule_room']
    h = reports['instructor_credits']

    with open(f'{folder}/table_course.md', 'w') as f:
        f.write(c)
    with open(f'{folder}/table_instructor.md', 'w') as f:
        f.write(n)
    with open(f'{folder}/table_room.md', 'w') as f:
        f.write(r)
    with open(f'{folder}/table_time.md', 'w') as f:
        f.write(t)

    print(ic)
    print(rc)
    display(Markdown(h))


    if (not instructor_conflicts) and (not room_conflicts):
        wbr = openpyxl.Workbook()
        wbr.remove(wbr.active)
        room_excel(wbr, df)
        wbr.save(Path(folder) / "schedule_room.xlsx")

        wbu = openpyxl.Workbook()
        wbu.remove(wbu.active)
        instructor_excel(wbu, df)
        wbu.save(Path(folder) / "schedule_instructor.xlsx")

    df.to_excel(Path(folder) / "schedule.xlsx", index=False)

